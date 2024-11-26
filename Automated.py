import os
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from solders.rpc.responses import GetTransactionResp
import layouts
import asyncio
import datetime
import websockets
import json
from solders.signature import Signature
from solana.rpc.async_api import AsyncClient
from datetime import datetime
from spl.token.constants import TOKEN_PROGRAM_ID, WRAPPED_SOL_MINT
from solana.rpc import types
from solana.rpc.types import TokenAccountOpts
from solders.pubkey import Pubkey
import sqlite3
import requests
from dotenv import dotenv_values
config = dotenv_values(".env")

class style():
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'


# wallet_address = "675kPX9MHTjS2zt1qfr1NYHuzeLXfQM9H24wFSUt1Mp8" #
seen_signatures = set()
Pool_raydium = "675kPX9MHTjS2zt1qfr1NYHuzeLXfQM9H24wFSUt1Mp8"
banangun_fees="4BBNEVRgrxVKv9f7pMNE788XM1tt379X9vNjpDH2KCL7"
solTrading_fees="HEPL5rTb6n1Ax6jt9z2XMPFJcDe9bSWvWQpsK7AMcbZg"
raydium_V4 = "5Q544fKrFoe6tsEbD7S8EmxGTJYAKtTVhAW5Q5pge4j1"


class TransactionProcessor:
    def __init__(self):
        self.queue = asyncio.Queue()
        self.wallet_address = None
        self.async_solana_client = AsyncClient(config["RPC_HTTPS_URL"])
        self.db_name = "automate.db"
        self.conn = sqlite3.connect(self.db_name)
        self.c = self.conn.cursor()
        self.init_db()
        self.income = 0
        self.outcome = 0
        self.fee = 0
        self.spent_sol = 0
        self.earned_sol = 0
        self.buys = 0
        self.sells = 0
        self.delta_sol=0
        self.delta_token=0
        self.delta_percentage=0
        self.first_buy_time = None
        self.last_sell_time = None
        self.last_trade = None
        self.time_period = 0
        self.contract = None
        self.scam_tokens = 0
        self.sol_balance = None
        self.solana_price = self.get_current_solana_price()
        self.tokenCreationTime = 0
        self.mint_decimal = None
        self.mint_address = None
        self.buy_period=0



    def init_db(self):
        self.c.execute('''
                    CREATE TABLE IF NOT EXISTS wallet_address (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        wallet_address TEXT UNIQUE
                    )
                ''')
        self.c.execute('''
                    CREATE TABLE IF NOT EXISTS token_accounts (
                        wallet_address_id INTEGER,
                        wallet_token_account TEXT,
                        block_time INTEGER,
                        FOREIGN KEY(wallet_address_id) REFERENCES wallet_address(id)
                    )
                ''')

        self.c.execute('''
                 CREATE TABLE IF NOT EXISTS pnl_info (
                     token_account TEXT PRIMARY KEY,
                     wallet_address_id INTEGER,
                     income REAL,
                     outcome REAL,
                     total_fee REAL,
                     spent_sol REAL,
                     earned_sol REAL,
                     delta_token REAL,
                     delta_sol REAL,
                     delta_percentage REAL,
                     buys INTEGER,
                     sells INTEGER,
                     last_trade TEXT,
                     time_period TEXT,
                     contract TEXT,
                     scam_tokens TEXT,
                     buy_period TEXT, 

                     FOREIGN KEY(wallet_address_id) REFERENCES wallet_address(id)
                 )
             ''')

        self.c.execute('''
               CREATE TABLE IF NOT EXISTS winning_wallets (
                   wallet_address_id INTEGER,
                   win_rate_7 REAL,
                   balance_change_7 REAL,
                   token_accounts_7 INTEGER,
                   win_rate_14 REAL,
                   balance_change_14 REAL,
                   token_accounts_14 INTEGER,
                   win_rate_30 REAL,
                   balance_change_30 REAL,
                   token_accounts_30 INTEGER,
                   win_rate_60 REAL,
                   balance_change_60 REAL,
                   token_accounts_60 INTEGER,
                   win_rate_90 REAL,
                   balance_change_90 REAL,
                   token_accounts_90 INTEGER,
                   FOREIGN KEY(wallet_address_id) REFERENCES wallet_address(id)
               )
           ''')
        self.conn.commit()

    async def get_token_accountsCount(self,wallet_address: Pubkey):
        owner = wallet_address
        opts = types.TokenAccountOpts(program_id=Pubkey.from_string("TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"))
        response = await self.async_solana_client.get_token_accounts_by_owner(owner, opts)
        return len(response.value)


    async def initialize(self):
        self.sol_balance = await self.getSOlBalance()



    """REPORT GENERATION FUNCTIONS"""
    async def getSOlBalance(self):
        pubkey = self.wallet_address
        response = await self.async_solana_client.get_balance(pubkey)
        balance = response.value /   10**9
        return balance

    def get_current_solana_price(self):
        url = "https://api.coingecko.com/api/v3/simple/price?ids=solana&vs_currencies=usd"
        response = requests.get(url)
        data = response.json()
        return data['solana']['usd'] if response.status_code ==   200 else None



    def store_win_rate(self, time_period, win_rate, balance_change, token_accounts):
        check_sql = 'SELECT 1 FROM winning_wallets WHERE wallet_address_id = ?'
        self.c.execute(check_sql, (self.wallet_address_id,))
        row_exists = self.c.fetchone() is not None
        time_period_suffix = f"_{time_period}"

        if row_exists:
            update_sql = f'''
                UPDATE winning_wallets
                SET win_rate{time_period_suffix} = ?, 
                    balance_change{time_period_suffix} = ?, 
                    token_accounts{time_period_suffix} = ?
                WHERE wallet_address_id = ?
            '''
            self.c.execute(update_sql, (win_rate, balance_change, token_accounts, self.wallet_address_id))
        else:

            insert_sql = f'''
                INSERT INTO winning_wallets (
                    wallet_address_id, 
                    win_rate_7, balance_change_7, token_accounts_7, 
                    win_rate_14, balance_change_14, token_accounts_14,
                    win_rate_30, balance_change_30, token_accounts_30,
                    win_rate_60, balance_change_60, token_accounts_60,
                    win_rate_90, balance_change_90, token_accounts_90
                ) VALUES (
                    ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL
                )
            '''
            self.c.execute(insert_sql, (self.wallet_address_id,))
            update_sql = f'''
                UPDATE winning_wallets
                SET win_rate{time_period_suffix} = ?, 
                    balance_change{time_period_suffix} = ?, 
                    token_accounts{time_period_suffix} = ?
                WHERE wallet_address_id = ?
            '''
            self.c.execute(update_sql, (win_rate, balance_change, token_accounts, self.wallet_address_id))

        self.conn.commit()

    def get_summary(self, time_period):
        query = f'''
            WITH Calculations AS (
              SELECT
                (SUM(CASE WHEN delta_sol > 0 THEN 1 ELSE 0 END) * 1.0 / COUNT(*)) * 100 AS WinRate,
                SUM(delta_sol) AS PnL_R,
                SUM(CASE WHEN delta_sol < 0 THEN delta_sol ELSE 0 END) AS PnL_Loss,
                (SUM(earned_sol) / NULLIF(SUM(spent_sol), 0) - 1) * 100 AS Balance_Change,
                COUNT(CASE WHEN scam_tokens = 1 THEN 1 END) AS ScamTokens,
                COUNT(token_account) AS TokenAccounts
               FROM pnl_info
          WHERE wallet_address_id = ?
            AND last_trade >= strftime('%s', 'now', '-{time_period} days')
        )
            SELECT
              *,
              '{time_period} days' AS TimePeriod
            FROM Calculations;
        '''

        self.c.execute(query, (self.get_wallet_address_id(self.wallet_address),))

        summary_result = self.c.fetchone()
        win_rate = summary_result[0]
        balance_change = summary_result[3]
        count_token_accounts=summary_result[5]
        self.store_win_rate(time_period, win_rate,balance_change,count_token_accounts)



        summary_data = {
            'SolBalance': self.sol_balance,
            'WalletAddress': str(self.wallet_address),
            'WinRate': summary_result[0],
            'PnL_R': summary_result[1],
            'PnL_Loss': summary_result[2],
            'Balance_Change': summary_result[3],
            'ScamTokens': summary_result[4],
            'TimePeriod': summary_result[6]  #
        }

        return summary_data
    def get_transactions(self, time_period):
        query = f'''
        SELECT *
        FROM pnl_info
        WHERE wallet_address_id = ?
          AND last_trade >= strftime('%s', 'now', '-{time_period} days');
        '''
        self.c.execute(query, (self.wallet_address_id,))
        results = self.c.fetchall()
        transactions_df = pd.DataFrame(results, columns=['token_account', 'wallet_address_id', 'income', 'outcome',
                                                         'total_fee', 'spent_sol', 'earned_sol', 'delta_token',
                                                         'delta_sol', 'delta_percentage', 'buys', 'sells',
                                                         'last_trade', 'time_period', 'contract', 'scam token','buy_period'])

        # Sort transactions by 'last_trade' in descending order
        transactions_df = transactions_df.sort_values(by='last_trade', ascending=False)

        return transactions_df

    async def generate_reports_for_time_periods(self, wallet,time_periods):
        await self.initialize()
        self.wallet_address_id = self.get_wallet_address_id(wallet)
        reports_folder = "automate_reports"
        summary_30_days = self.get_summary(30)
        win_rate_30_days = summary_30_days['WinRate'] if summary_30_days else 'Unknown'
        if not os.path.exists(reports_folder):
            os.makedirs(reports_folder)
        wallet_folder = os.path.join(reports_folder, f"{win_rate_30_days}_days_{wallet}")
        if not os.path.exists(wallet_folder):
            os.makedirs(wallet_folder)

        for time_period in time_periods:
            summary = self.get_summary(time_period)
            transactions = self.get_transactions(time_period)
            if summary:
                file_name = os.path.join(wallet_folder, f"{self.wallet_address}_{time_period}_days.xlsx")
                self.export_to_excel(summary, transactions, file_name)
                print(f"Exported summary and transactions for {time_period} days to {file_name}")
            else:
                print(f"No summary found for {time_period} days.")

    def export_to_excel(self,summary, transactions, file_name):
        sol_current_price = self.solana_price
        summary['Current_Sol_Price'] = sol_current_price
        summary['ID'] = self.wallet_address_id
        try:
            summary['Profit_USD'] = abs(summary['PnL_R']) * sol_current_price
            summary['Loss_USD'] = abs(summary['PnL_Loss']) * sol_current_price
        except TypeError:
            summary['Profit_USD'] = 0
            summary['Loss_USD'] = 0

        summary_df = pd.DataFrame([summary], columns=['ID','WalletAddress', 'SolBalance','Current_Sol_Price', 'WinRate', 'PnL_R', 'PnL_Loss',
                                                      'Balance_Change', 'ScamTokens', 'Profit_USD',
                                                      'Loss_USD', 'TimePeriod'])

        transactions_df = pd.DataFrame(transactions, columns=['token_account','wallet_address_id', 'income', 'outcome',
                                                              'total_fee', 'spent_sol', 'earned_sol', 'delta_token',
                                                              'delta_sol', 'delta_percentage', 'buys', 'sells',
                                                              'last_trade', 'time_period','buy_period', 'contract', 'scam token'])
        transactions_df.drop(columns=['wallet_address_id'], inplace=True)
        transactions_df['last_trade'] = transactions_df['last_trade'].apply(lambda x: datetime.fromtimestamp(int(x)).strftime('%d.%m.%Y'))


        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary and Transactions', index=False, startrow=0)
            row_to_start = len(summary_df) + 2
            transactions_df.to_excel(writer, sheet_name='Summary and Transactions', index=False, startrow=row_to_start)
            workbook = writer.book
            worksheet = writer.sheets['Summary and Transactions']

            for row in worksheet.iter_rows(min_row=row_to_start + 2, max_col=worksheet.max_column):
                for cell in row:
                    if cell.column == 1:  # 'token_account' column
                        cell.hyperlink = f'https://solscan.io/account/{cell.value}#splTransfer'
                        cell.value = 'View Solscan'
                        cell.font = Font(underline='single')
                    elif cell.column == 15:  # 'contract' column
                        cell.hyperlink = f'https://dexscreener.com/solana/{cell.value}?maker={self.wallet_address}'
                        'https://dexscreener.com/solana/3zcoadmvqtx3itfthwr946nhznqh92eq9hdhhjtgp6as?maker=3uij3uDg5pLBBxQw6hUXxqw6uEwCQGX7rM8PZb7ofH9e'
                        cell.value = 'View Dexscreener'
                        cell.font = Font(underline='single')


            # Define fills for conditional formatting
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            brown_fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
            gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

            # Apply initial styling
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.row == 1:
                        cell.font = Font(bold=True)

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

            for idx, row in enumerate(transactions_df.itertuples(index=False), start=row_to_start + 2):
                outcome = row[2]
                income=row[1]
                delta_percentage = row[8]
                time_period = row[12]
                buys = row[9]


                if round(outcome,1) > round(income,1):
                    worksheet.cell(row=idx, column=3).fill = yellow_fill


                if delta_percentage ==-100:

                    worksheet.cell(row=idx, column=9).fill = brown_fill

                if pd.to_timedelta(time_period) < pd.Timedelta(minutes=1):
                    worksheet.cell(row=idx, column=13).fill = yellow_fill

                if buys > 3:
                    worksheet.cell(row=idx, column=10).fill = yellow_fill


            for idx, row in enumerate(transactions_df.itertuples(index=False), start=row_to_start + 2):
                # For delta_sol
                if row[7] < 0:
                    worksheet.cell(row=idx, column=8).fill = red_fill
                elif row[7] > 0:
                    worksheet.cell(row=idx, column=8).fill = green_fill

                if row[7] < 0 and row[8] != -100:
                    worksheet.cell(row=idx, column=9).fill = red_fill
                elif row[8] > 0:
                    worksheet.cell(row=idx, column=9).fill = green_fill


            for idx, row in enumerate(summary_df.itertuples(index=False), start=1):
                if row[9] > row[10] and row[4]>=50:  # If Profit_USD > Loss_USD
                    worksheet.cell(row=idx, column=10).fill = gold_fill
                    worksheet.cell(row=idx, column=5).fill = gold_fill
                elif row[9] < row[10] and row[4]<50:  # If   Loss_USD> Profit_USD
                    worksheet.cell(row=idx, column=10).fill = red_fill
                    worksheet.cell(row=idx, column=5).fill = red_fill
                elif row[9] < row[10] and row[4] == 50:
                    worksheet.cell(row=idx, column=10).fill = red_fill
                    worksheet.cell(row=idx, column=5).fill = gold_fill
                else:
                    worksheet.cell(row=idx, column=11).fill = red_fill
                    worksheet.cell(row=idx, column=5).fill = red_fill
            workbook.save(file_name)




    def get_wallet_address_id(self,wallet_address):
        self.c.execute('SELECT id FROM wallet_address WHERE wallet_address = ?', (str(wallet_address),))
        result = self.c.fetchone()
        if result:
            return result[0]
        else:
            self.c.execute('INSERT INTO wallet_address (wallet_address) VALUES (?)', (str(wallet_address),))
            self.conn.commit()
            return self.c.lastrowid

    def convert_unix_to_date(self,unix_timestamp):
        """
        Convert a Unix timestamp to a datetime object.
        """
        timestamp_str = str(unix_timestamp)
        if len(timestamp_str) > 10:
            return datetime.fromtimestamp(round(unix_timestamp / 1000))

        return datetime.fromtimestamp(unix_timestamp)


    async def update_token_account(self, wallet_address, wallet_token_account, block_time):
        try:
            wallet_address_id = self.get_wallet_address_id(str(wallet_address))
            if wallet_address_id is None:
                print("Wallet address not found in the database.")
                return

            # Convert the Pubkey to a string
            wallet_token_account_str = str(wallet_token_account)  # Adjust this line as necessary

            # Check if the token account already exists for this wallet address
            self.c.execute('SELECT * FROM token_accounts WHERE wallet_address_id = ? AND wallet_token_account = ?',
                           (wallet_address_id, wallet_token_account_str))
            result = self.c.fetchone()

            if not result:
                # Insert a new token account only if it does not exist
                self.c.execute(
                    'INSERT INTO token_accounts (wallet_address_id, wallet_token_account, block_time) VALUES (?, ?, ?)',
                    (wallet_address_id, wallet_token_account_str, block_time))
                # Commit the changes
                self.conn.commit()
                print(f"{style.CYAN}New token account added for wallet address: {str(wallet_address)}, token account: {wallet_token_account_str}",style.RESET)
                print("Calculating and updating pnl")
                print(wallet_token_account)
                await self.process_token_account(wallet_token_account,wallet_address_id)
            else:
                print("Account already exists")

        except Exception as e:
            print(f"Error updating token account: {e}")
            self.conn.rollback()



    def token_account_exists(self,wallet_address, wallet_token_account):
        wallet_address_id = self.get_wallet_address_id(str(wallet_address))
        if wallet_address_id is None:
            print("Wallet address not found in the database.")
            return False


        self.c.execute('SELECT * FROM token_accounts WHERE wallet_address_id = ? AND wallet_token_account = ?',
                  (wallet_address_id, str(wallet_token_account)))
        result = self.c.fetchone()

        if result:
            return True
        else:
            return False

    async def get_new_token_accounts(self,wallet_address: Pubkey):
        owner = self.wallet_address
        opts = TokenAccountOpts(
            program_id=Pubkey.from_string("TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA")
        )
        response = await self.async_solana_client.get_token_accounts_by_owner(owner, opts)
        number_tokenAccounts = len(response.value)
        all_token_accounts = response.value
        new_token_accounts = []
        for token_account in all_token_accounts:
            sig= await self.async_solana_client.get_signatures_for_address(token_account.pubkey, limit=100)
            last_signature = await self.async_solana_client.get_transaction(sig.value[-1].signature, encoding="jsonParsed",
                                                       max_supported_transaction_version=0)

            block_time=last_signature.value.block_time
            account_exist = self.token_account_exists(wallet_address, str(token_account.pubkey))
            if not account_exist:
                new_token_accounts.append(token_account)
                await self.update_token_account(wallet_address, token_account.pubkey, block_time)
            else:
                print("Account already exists")
        print(f"{style.GREEN}Added all token account of {wallet_address} to the database",style.RESET)
        print("------------Generating Report---------------")
        reports_folder = "automate_processing_reports"
        summary_30_days = self.get_summary(30)
        win_rate_30_days = summary_30_days['WinRate'] if summary_30_days else 'Unknown'

        for time_period in [90, 60, 30, 14, 7]:
            summary = self.get_summary(time_period)
            transactions = self.get_transactions(time_period)

            if summary:
                # Create a new folder for each wallet address
                wallet_folder = os.path.join(reports_folder, f"{win_rate_30_days}_{self.wallet_address}")
                if not os.path.exists(wallet_folder):
                    os.makedirs(wallet_folder)
                file_name = os.path.join(wallet_folder, f"{self.wallet_address}_{time_period}_days.xlsx")
                self.export_to_excel(summary, transactions, file_name)
                print(f"Exported summary and transactions for {time_period} days to {file_name}")
            else:
                print(f"No summary found for {time_period} days.")

        return new_token_accounts

    ### PNL INFO####
    def get_token_data(self,decimals):
        for token_balances in decimals:
            if token_balances.owner == self.wallet_address and token_balances.mint != Pubkey.from_string(WRAPPED_SOL_MINT):
                token_contract = token_balances.mint
                token_decimal = token_balances.ui_token_amount.decimals
                return token_contract, token_decimal
    async def transactionType(self,Account:str):
        data_response = await self.async_solana_client.get_account_info(Pubkey.from_string(Account))
        data = data_response.value.data
        parsed_data = layouts.SPL_ACCOUNT_LAYOUT.parse(data)
        mint = Pubkey.from_bytes(parsed_data.mint)
        if mint == WRAPPED_SOL_MINT:
            return mint
        return mint

    async def transactionDetails(self , txn: GetTransactionResp,transaction_array:list):
          transaction= txn
          information_array = transaction_array
          block_time = transaction.value.block_time
          txn_fee = transaction.value.transaction.meta.fee
          mint_decimal = self.mint_decimal
          mint_address=self.mint_address
          pre_tokenBalance = transaction.value.transaction.meta.pre_token_balances
          post_tokenBalance = transaction.value.transaction.meta.post_token_balances
          tokenAmount_Sold = pre_tokenBalance[-1].ui_token_amount.ui_amount
          tokenAuthority = pre_tokenBalance[-1].owner
          jupyter_transaction={}

          print(style.RED + "Length of inforamtion araay", len(information_array), style.RESET)

          if len(information_array)==2:

              if len(pre_tokenBalance) > 0:

                if pre_tokenBalance[0].mint != Pubkey.from_string("So11111111111111111111111111111111111111112"):
                    mint_address = pre_tokenBalance[0].mint
                    mint_decimal = pre_tokenBalance[0].ui_token_amount.decimals
                    print(mint_decimal,mint_address)
                    tokenAmount_Bought = post_tokenBalance[0].ui_token_amount.ui_amount
                else:
                    mint_address = pre_tokenBalance[1].mint
                    mint_decimal = pre_tokenBalance[1].ui_token_amount.decimals
                    print(mint_decimal,mint_address)
              else:
                print("This is a transfer Continue")

              try:

                first_info = information_array[0]
                second_info = information_array[1]
                first_authority = first_info['authority']
                transfer_type= await self.transactionType(second_info['source'])
                token_sold = int(first_info['amount']) / 10 ** mint_decimal
                sol_sold = int(second_info['amount']) / 10 ** 9
                sol_spent = int(first_info['amount']) / 10 ** 9
                token_bought = int(second_info['amount']) / 10 ** mint_decimal

                if first_authority== str(self.wallet_address) and str(transfer_type) != str(WRAPPED_SOL_MINT):
                  self.update_buy(token_bought, txn_fee, block_time)
                  self.spent_sol += sol_spent
                  self.contract = mint_address
                  self.last_trade = block_time
                  print(f"{style.GREEN}BUY {sol_spent} SOL {style.RESET} -FOR  {token_bought} TokenBought= {mint_address}")
                else:
                  self.update_sell(token_sold, txn_fee, block_time)
                  self.earned_sol += sol_sold
                  self.contract = mint_address
                  self.last_trade = block_time
                  print(f"{style.RED}SELL {int(token_sold)} Token {style.RESET} -FOR  {sol_sold} SOL TokenSold= {mint_address}")


              except Exception as e:
                  print("Error",e)

          else:

              try:

                  jupyter_transaction['first'] = information_array[0]
                  jupyter_transaction['last'] = information_array[-1]

                  if str(jupyter_transaction['last']['mint']) != str(WRAPPED_SOL_MINT) and str(jupyter_transaction['last']['mint']) != "EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v":

                      print(style.YELLOW + "This is jupyter BUYYY", style.RESET)
                      from_wallet = jupyter_transaction['first']['authority']
                      to_wallet = jupyter_transaction['last']['authority']
                      buy_amount = jupyter_transaction['first']['tokenAmount']['uiAmount']
                      mint = jupyter_transaction['last']['mint']
                      tokenAmount_Bought = jupyter_transaction['last']['tokenAmount']['uiAmount']

                      self.update_buy(tokenAmount_Bought, txn_fee, block_time)
                      self.spent_sol += buy_amount
                      self.contract = mint_address
                      self.last_trade = block_time
                      print(f"Buy  amount {buy_amount}SOL for Token= {mint} Token Amount Bought= {tokenAmount_Bought}")
                      jupyter_transaction={}
                  else:
                      print(style.BLUE + "This is jupyter SELL", style.RESET)
                      from_wallet = jupyter_transaction['first']['authority']
                      to_wallet = jupyter_transaction['last']['authority']
                      tokenAmount_Sold = jupyter_transaction['first']['tokenAmount']['uiAmount']
                      mint = jupyter_transaction['first']['mint']
                      Soll_sell_amount = jupyter_transaction['last']['tokenAmount']['uiAmount']
                      self.update_sell(tokenAmount_Sold, txn_fee, block_time)
                      self.earned_sol += Soll_sell_amount
                      self.contract = mint
                      self.last_trade = block_time
                      print(f"Sell  amount {tokenAmount_Sold} Token= {mint} for SOL= {Soll_sell_amount}")
                      jupyter_transaction = {}
              except Exception as e:
                    print("Jupyter Error", e)


    async def process_token_account(self, token_account: Pubkey,wallet_address_id: int):
        transaction_data_dict = {}
        error_count = 0
        self.reset_variables()
        token_account_str = str(token_account)

        # Directly process the transactions for the given token account
        sig = await self.async_solana_client.get_signatures_for_address(token_account, limit=500)
        valid=0
        # information_array = []
        for signature in reversed(sig.value):
            if signature.err == None:
                transaction = await self.async_solana_client.get_transaction(signature.signature, encoding="jsonParsed",
                                                                             max_supported_transaction_version=0)
                txn_fee = transaction.value.transaction.meta.fee
                instruction_list = transaction.value.transaction.meta.inner_instructions
                account_signer = transaction.value.transaction.transaction.message.account_keys[0].pubkey
                decimals = transaction.value.transaction.meta.post_token_balances
                information_array = []

                print(style.RED,signature.signature,style.RESET)
                for ui_inner_instructions in instruction_list:
                    for txn_instructions in ui_inner_instructions.instructions:
                        if txn_instructions.program_id == TOKEN_PROGRAM_ID:
                            txn_information = txn_instructions.parsed['info']
                            if 'destination' in txn_information:
                                information_array.append(txn_information)


                if account_signer == self.wallet_address:

                   valid+=1

                try:
                    if valid>0:
                        await  self.transactionDetails(transaction,information_array)

                        print(style.CYAN,"Transaction Details Added",style.RESET)
                        valid=0
                    else:
                        pass
                except Exception as e:
                    print(e, signature.signature)
                    print(style.RED,"Error Adding Transaction Details",style.RESET)
                    continue

        await self.calculate_deltas()
        self.print_summary()

        await self.fill_pnl_info_table(token_account_str,wallet_address_id)




    async def pair_createdTime(self,token_traded):
        url = f'https://api.dexscreener.com/latest/dex/tokens/{token_traded}'
        response = requests.get(url)
        data = response.json()
        if data['pairs'] is not None:
            return data['pairs'][0]['pairCreatedAt']
        return 0


    def calculate_time_difference(self,unix_timestamp1, unix_timestamp2):
        """
        Calculate the difference between two Unix timestamps and return it in a human-readable format.
        """
        date1 = self.convert_unix_to_date(unix_timestamp1)
        date2 = self.convert_unix_to_date(unix_timestamp2)

        # Calculate the difference between the two dates
        difference = date2 - date1

        # Calculate the difference in hours, minutes, and seconds
        hours, remainder = divmod(difference.total_seconds(), 3600)
        minutes, seconds = divmod(remainder, 60)

        # Format the output based on the difference
        if hours > 0:
            return f"{int(hours)}h {int(minutes)}m {int(seconds)}s"
        elif minutes > 0:
            return f"{int(minutes)}m {int(seconds)}s"
        else:
            return f"{int(seconds)}s"
    def update_buy(self, amount, fee, block_time):

        self.income += amount
        self.fee += fee
        self.buys +=  1
        if self.first_buy_time is None:
            self.first_buy_time = block_time
            if self.tokenCreationTime != 0:
                self.buy_period= self.calculate_time_difference(self.first_buy_time, self.tokenCreationTime)
            else:
                self.buy_period="Unknown"



    def update_sell(self, amount, fee, block_time):
        self.outcome += amount
        self.fee += fee
        self.sells +=  1
        self.last_sell_time = block_time
        if self.last_trade is None or block_time > self.last_trade:
            self.last_trade = block_time
    def print_summary(self):
        print(f"Income: {self.income}")
        print(f"Outcome: {self.outcome}")
        print(f"Total Fee: {self.fee/10**9}")
        print(f"Spent SOL: {self.spent_sol}")
        print(f"Earned SOL: {self.earned_sol}")
        print(f"Delta Token: {self.delta_token}")
        print(f"Delta SOL: {self.delta_sol}")
        print(f"Delta Percentage: {self.delta_percentage}%")
        print(f"Buys: {self.buys}")
        print(f"Sells: {self.sells}")
        print(f"Last Trade:{self.convert_unix_to_date(self.last_trade)}")###Check if this conversion is the problem
        print(f"Time Period: {self.time_period}")
        print(f"Contract: {self.contract}")
        print(f"Scam Tokens: {self.scam_tokens}")
        print(self.tokenCreationTime, self.first_buy_time)
        print("Buy Period: ",self.buy_period)



    def reset_variables(self):
        self.income = 0
        self.outcome = 0
        self.fee = 0
        self.spent_sol = 0
        self.earned_sol = 0
        self.buys = 0
        self.sells = 0
        self.first_buy_time = None
        self.last_sell_time = None
        self.last_trade = None
        self.time_period = 0
        self.contract = None
        self.scam_tokens = 0
        self.buy_period=0
        self.tokenCreationTime = 0
        self.mint_decimal=None


    async def getToken_SolAmount(self,):
        token_address = str(self.contract)
        url = f'https://api.dexscreener.com/latest/dex/tokens/{token_address}'
        response = requests.get(url)
        data = response.json()
        if data['pairs'] is not None:
            token_price_usd = float(data['pairs'][0]['priceUsd'])
            wallet_amount= self.income
            Worth_wallet_amount = token_price_usd * wallet_amount
            worth_in_solana= Worth_wallet_amount / self.solana_price
            return worth_in_solana
        else:
            self.scam_tokens = 1
            return 0





    async def calculate_deltas(self):

        if self.sells >=1:
            self.delta_token = self.income - self.outcome
            self.delta_sol = self.earned_sol - self.spent_sol
            self.delta_percentage = (self.delta_sol / self.spent_sol) * 100 if self.spent_sol != 0 else 0
        else:
            self.delta_token= self.income
            self.delta_sol = self.earned_sol-self.spent_sol
            self.delta_percentage= -100
        self.tokenCreationTime = await self.pair_createdTime(self.contract)
        if self.tokenCreationTime == 0:
            self.buy_period = "Unknown"

        else:
            self.buy_period = self.calculate_time_difference(self.tokenCreationTime, self.first_buy_time)

        # Calculate the time period
        if self.first_buy_time and self.last_sell_time:
            time_difference = self.last_sell_time - self.first_buy_time
            self.time_period = self.calculate_time_difference(self.first_buy_time, self.last_sell_time)
        elif self.first_buy_time and not self.last_sell_time:
            self.time_period = 0  # No sell, indicating a potential scam
            self.scam_tokens = 1


    async def fill_pnl_info_table(self, token_account,wallet_address_id):
        """
        Insert a new PNL info only if the transaction is not yet in the database.
        """

        fields = [

            ('token_account', token_account),
            ('income', self.income),
            ('outcome', self.outcome),
            ('fee', self.fee),
            ('spent_sol', self.spent_sol),
            ('earned_sol', self.earned_sol),
            ('delta_token', self.delta_token),
            ('delta_sol', self.delta_sol),
            ('buys', self.buys),
            ('sells', self.sells),
            ('time_period', self.time_period),
            ('contract', self.contract),
            ('scam_tokens', self.scam_tokens),
            ('wallet_address_id', wallet_address_id),
            ('last_trade', self.last_trade),
            ('buy_period',self.buy_period)
        ]

        none_fields = [field_name for field_name, field_value in fields if field_value is None]
        if none_fields:
            print(f"One or more fields are None: {', '.join(none_fields)}. Skipping the operation.")
            return
        # Prepare the SQL INSERT statement
        insert_sql = '''
                INSERT INTO pnl_info (
                    wallet_address_id, token_account, income, outcome, total_fee, spent_sol, earned_sol,
                    delta_token, delta_sol, delta_percentage, buys, sells, last_trade,
                    time_period, contract, scam_tokens,buy_period
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?)
            '''

        if any(value is None for value in
               [token_account, self.income, self.outcome, self.fee, self.spent_sol, self.earned_sol, self.delta_token,
                self.delta_sol, self.buys, self.sells, self.time_period, self.contract, self.scam_tokens,
                wallet_address_id, self.last_trade, self.calculate_time_difference(self.first_buy_time, self.tokenCreationTime) if self.tokenCreationTime != 0 or self.first_buy_time!= None else 0]):
            print("One or more fields are None. Skipping the operation.")
            return

        # Check if the transaction already exists in the database
        check_sql = '''
                SELECT   1 FROM pnl_info
                WHERE wallet_address_id = ? AND last_trade = ?
            '''
        cursor = self.conn.cursor()
        cursor.execute(check_sql, (wallet_address_id, self.last_trade))
        if cursor.fetchone() is not None:
            print("Transaction already exists in the database. Skipping the operation.")
            return

        insert_data = (
            wallet_address_id,
            token_account,  # Use the token_account parameter
            self.income,
            self.outcome,
            self.fee / 10 ** 9,
            self.spent_sol,
            self.earned_sol,
            self.delta_token,
            self.delta_sol,
            self.delta_percentage,
            self.buys,
            self.sells,
            self.last_trade,
            self.time_period,
            str(self.contract),
            self.scam_tokens,
            self.buy_period
        )

        # Insert the new row
        cursor.execute(insert_sql, insert_data)
        self.conn.commit()
        print("PNL info successfully inserted into the database.")

    async def process_transactions(self):
        while True:
            signature = await self.queue.get()
            try:
                transaction = await self.async_solana_client.get_transaction(signature, encoding="jsonParsed",
                                                                             max_supported_transaction_version=0)
                instruction_list = transaction.value.transaction.meta.inner_instructions
                accounts = transaction.value.transaction.transaction.message.account_keys
                if accounts[-1].pubkey == Pubkey.from_string(Pool_raydium):
                    self.wallet_address = accounts[0].pubkey
                    await self.initialize()

                    self.wallet_address_id = self.get_wallet_address_id(self.wallet_address)
                    num_tokenAccounts = await self.get_token_accountsCount(self.wallet_address)
                    print("Number of token Account", self.wallet_address, num_tokenAccounts)
                    if num_tokenAccounts in range(1, int(config['max_token_accounts']) + 1):
                        await self.get_new_token_accounts(self.wallet_address)


            except Exception as e:
                print(f"Failed to process transaction {signature}: {e}")
                pass
            finally:
                self.queue.task_done()

    async def enqueue_transaction(self, signature):
        await self.queue.put(signature)


async def run():
    processor = TransactionProcessor()
    asyncio.create_task(processor.process_transactions())
    # await asyncio.gather(
    #     # processor.process_transactions(),
    #     processor.generate_reports_for_time_periods("9KupEacYuc5Pt7A8nhYnFLnwFFsVR9othS7azjV8rqzP",[365]),
    #     # processor.get_summary(1),
    #     processor.initialize()
    # )

    while True:  # Loop for reconnection attempts
        try:

            uri = "wss://mainnet.helius-rpc.com/?api-key=463c5e26-e304-4db4-afc3-16d6c10fe29e"
            async with websockets.connect(uri, ping_timeout=30) as websocket:
                await websocket.send(json.dumps({
                    "jsonrpc": "2.0",
                    "id": 1,
                    "method": "logsSubscribe",
                    "params": [
                        {"mentions": [config['automated_wallet_address']]},
                        {"commitment": "finalized"}
                    ]
                }))

                first_resp = await websocket.recv()
                response_dict = json.loads(first_resp)
                if 'result' in response_dict:
                    print("Subscription successful. Subscription ID: ", response_dict['result'])
                    async for response in websocket:
                        response_dict = json.loads(response)
                        if response_dict['params']['result']['value']['err'] is None:
                            signature = response_dict['params']['result']['value']['signature']
                            if signature not in seen_signatures:
                                seen_signatures.add(signature)
                                await processor.enqueue_transaction(Signature.from_string(signature))

        except websockets.exceptions.ConnectionClosedError as e:
            print(f"Connection closed with error: {e}. Reconnecting in 5 seconds...")
            await asyncio.sleep(5)
        except Exception as e:
            print(f"An unexpected error occurred: {e}. Attempting reconnect in 5 seconds...")
            await asyncio.sleep(5)


asyncio.run(run())